"""
Complete business-plan Excel export (VIPA-aligned).

Entry point: generate_excel_report(plan_data, output_path)
plan_data shape (from GET plan + results or worker assembly):
  {
    "inputs": { ... PlanInputs ... },
    "results": { ... PlanResults ... },
    "yearly": [ { "year", "revenue", "cogs", ... }, ... ],  # optional
    "title": str,
    "plan_id": str,
  }
"""

from __future__ import annotations

import math
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from bp_calc.engine import HORIZON, calculate_plan
from bp_calc.projections import compute_yearly_pl_breakdown
from bp_calc.capex import annual_depreciation_schedule, total_capex
from bp_calc.loan import build_amortization_schedule
from bp_schema.liasse import PlanInputs, PlanResults

# —— Style constants ——
FONT_NAME = "Arial"
FONT_SIZE = 10
COL_YEAR_WIDTH = 14
NUM_FMT = '#,##0;(#,##0);"-"'
PCT_FMT = "0.0%"

CLR_HEADER = "003366"
CLR_SUBTOTAL = "D3D3D3"
CLR_TOTAL = "336699"
CLR_INPUT = "0000FF"
CLR_KPI_VAN = "E8F4FC"
CLR_PROD = "E8F5E9"
CLR_CHARGE = "FFEBEE"
CLR_RESULT = "E3F2FD"
CLR_DRCI = "FFF9C4"

YEAR_LABELS = [f"An {i + 1}" for i in range(HORIZON)]
CA_FACTORS = [0.8, 0.9, 1.0, 1.1, 1.2]
COST_FACTORS = [0.8, 0.9, 1.0, 1.1, 1.2]

_thin = Side(style="thin", color="CCCCCC")
BORDER_ALL = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _font(*, bold=False, color="000000", size=FONT_SIZE) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _align(h="left", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


def _series(obj: Any, key: str, default: float = 0.0) -> list[float]:
    if obj is None:
        return [default] * HORIZON
    if isinstance(obj, dict):
        years = obj.get("years") or obj.get(key)
    else:
        years = getattr(obj, "years", None)
    if years is None:
        return [default] * HORIZON
    out = [float(v) for v in years[:HORIZON]]
    while len(out) < HORIZON:
        out.append(default)
    return out


def _safe_div(a: float, b: float) -> float:
    return a / b if b else 0.0


@dataclass
class ReportContext:
    inputs: PlanInputs
    results: PlanResults
    yearly: list[dict[str, Any]]
    title: str
    plan_id: str

    @property
    def company(self) -> str:
        return (self.inputs.company.name or "Projet").strip()

    @property
    def ind(self):
        return self.results.indicators

    def rev(self) -> list[float]:
        return _series(self.results.revenue, "revenue")

    def net(self) -> list[float]:
        return _series(self.results.netProfit, "netProfit")

    def dep(self) -> list[float]:
        return _series(self.results.depreciation, "depreciation")

    def interest(self) -> list[float]:
        return _series(self.results.interestExpense, "interest")

    def principal(self) -> list[float]:
        return _series(self.results.principalRepayment, "principal")

    def bfr_var(self) -> list[float]:
        return _series(self.results.bfrVariation, "bfrVariation")

    def ocf(self) -> list[float]:
        return _series(self.results.operatingCashFlow, "operatingCashFlow")

    def cum_treasury(self) -> list[float]:
        return _series(self.results.cumulativeTreasury, "cumulativeTreasury")

    def purchase_mp(self) -> list[float]:
        return _series(self.results.purchaseValueMP, "purchaseValueMP")

    def mkt(self) -> list[float]:
        return _series(self.results.marketingExpense, "marketing")

    def dist(self) -> list[float]:
        return _series(self.results.distributionExpense, "distribution")


def build_plan_data(
    *,
    inputs: PlanInputs | dict,
    results: PlanResults | dict,
    plan_id: str = "",
    title: str = "",
    yearly: list[dict] | None = None,
) -> dict:
    inp = PlanInputs.model_validate(inputs) if isinstance(inputs, dict) else inputs
    res = PlanResults.model_validate(results) if isinstance(results, dict) else results
    yr = yearly
    if yr is None:
        _, yr = compute_yearly_pl_breakdown(inp)
    return {
        "inputs": inp.model_dump(),
        "results": res.model_dump(),
        "yearly": yr,
        "plan_id": plan_id,
        "title": title or f"Business Plan — {inp.company.name}",
    }


def _parse_plan_data(plan_data: dict) -> ReportContext:
    inputs = PlanInputs.model_validate(plan_data.get("inputs") or {})
    results = PlanResults.model_validate(plan_data.get("results") or {})
    yearly = plan_data.get("yearly")
    if not yearly:
        _, yearly = compute_yearly_pl_breakdown(inputs)
    return ReportContext(
        inputs=inputs,
        results=results,
        yearly=list(yearly or []),
        title=str(plan_data.get("title") or f"Business Plan — {inputs.company.name}"),
        plan_id=str(plan_data.get("plan_id") or ""),
    )


class _SheetWriter:
    def __init__(self, ws: Worksheet):
        self.ws = ws
        self.ws.sheet_view.showGridLines = True

    def setup_sheet(self) -> None:
        self.ws.sheet_properties.pageSetUpPr.fitToPage = True
        for c in range(2, 2 + HORIZON):
            self.ws.column_dimensions[get_column_letter(c)].width = COL_YEAR_WIDTH
        self.ws.column_dimensions["A"].width = 32

    def cell(self, row: int, col: int, value: Any = None, *, style: str = "data") -> None:
        c = self.ws.cell(row=row, column=col, value=value)
        c.font = _font()
        c.border = BORDER_ALL
        c.alignment = _align("right" if col > 1 and isinstance(value, (int, float)) else "left")
        if style == "header":
            c.fill = _fill(CLR_HEADER)
            c.font = _font(bold=True, color="FFFFFF")
            c.alignment = _align("center")
        elif style == "subtotal":
            c.fill = _fill(CLR_SUBTOTAL)
            c.font = _font(bold=True)
        elif style == "total":
            c.fill = _fill(CLR_TOTAL)
            c.font = _font(bold=True, color="FFFFFF")
        elif style == "input":
            c.font = _font(color=CLR_INPUT)
        elif style == "label":
            c.alignment = _align("left")
        if isinstance(value, (int, float)) and style not in ("header",):
            c.number_format = NUM_FMT
        return c

    def freeze(self, cell: str = "B2") -> None:
        self.ws.freeze_panes = cell

    def write_row(
        self,
        row: int,
        labels: list[Any],
        *,
        start_col: int = 1,
        row_style: str = "data",
    ) -> None:
        for i, val in enumerate(labels):
            st = row_style if i == 0 and row_style != "data" else ("label" if i == 0 else "data")
            self.cell(row, start_col + i, val, style=st if st != "data" or i > 0 else "label")

    def write_year_header(self, row: int, label: str = "Poste") -> int:
        self.cell(row, 1, label, style="header")
        for i, yl in enumerate(YEAR_LABELS):
            self.cell(row, 2 + i, yl, style="header")
        return row + 1

    def write_year_series(
        self,
        row: int,
        label: str,
        values: list[float],
        *,
        row_style: str = "data",
        pct_row: bool = False,
    ) -> int:
        self.cell(row, 1, label, style="label" if row_style == "data" else row_style)
        for i, v in enumerate(values[:HORIZON]):
            self.cell(row, 2 + i, v if not pct_row else v, style=row_style)
            if pct_row:
                self.ws.cell(row, 2 + i).number_format = PCT_FMT
        return row + 1


def _product_names(ctx: ReportContext) -> list[str]:
    raw = ctx.inputs.model_dump()
    products = raw.get("products") if isinstance(raw, dict) else None
    if isinstance(products, dict):
        catalog = products.get("catalog") or []
        names = []
        for item in catalog:
            if isinstance(item, dict):
                n = (item.get("name") or item.get("label") or "").strip()
                if n:
                    names.append(n)
        if names:
            return names[:12]
    return ["Produit principal"]


def _categorize_investments(ctx: ReportContext) -> dict[str, list[tuple[str, float, int, float]]]:
    """category -> [(name, amount, life, rate)]"""
    buckets: dict[str, list] = {
        "incorporel": [],
        "agencement": [],
        "industriel": [],
        "transport": [],
        "bureau": [],
        "preliminaires": [],
    }
    keywords = {
        "transport": ("fourgon", "véhicule", "vehicule", "berlingo", "jumper", "transport"),
        "bureau": ("bureau", "ordinateur", "mobilier", "administratif"),
        "agencement": ("agencement", "climatisation", "électricité", "electricite", "sécurité", "portes"),
        "preliminaires": ("frais préliminaire", "publicité", "formation", "constitution", "étude"),
    }
    for eq in ctx.inputs.investments.equipment:
        if eq.cost <= 0 and not eq.name.strip():
            continue
        name_l = eq.name.lower()
        cat = "industriel"
        if eq.assetType == "intangible":
            cat = "incorporel"
        else:
            for bucket, kws in keywords.items():
                if any(k in name_l for k in kws):
                    cat = bucket
                    break
        rate = 1.0 / max(1, eq.usefulLifeYears)
        buckets[cat].append((eq.name, eq.cost, eq.usefulLifeYears, rate))
    for line in ctx.inputs.investments.intangible:
        if line.amount > 0:
            buckets["incorporel"].append(
                (line.label, line.amount, line.usefulLifeYears, 1.0 / max(1, line.usefulLifeYears))
            )
    for line in ctx.inputs.investments.tangible:
        if line.amount > 0:
            buckets["industriel"].append(
                (line.label, line.amount, line.usefulLifeYears, 1.0 / max(1, line.usefulLifeYears))
            )
    return buckets


def _loan_annual_rows(ctx: ReportContext) -> list[dict[str, float]]:
    fin = ctx.inputs.financing
    amount = fin.loan.amount or ctx.results.totalInvestment * fin.debtRatio
    periods = build_amortization_schedule(
        amount,
        fin.loan.rate,
        fin.loan.years,
        fin.loan.graceMonthsPrincipal,
        frequency="quarterly",
    )
    rows = []
    opening = amount
    for y in range(HORIZON):
        interest = ctx.interest()[y]
        principal = ctx.principal()[y]
        closing = max(0.0, opening - principal)
        rows.append(
            {
                "opening": opening,
                "interest": interest,
                "principal": principal,
                "service": interest + principal,
                "closing": closing,
            }
        )
        opening = closing
    return rows


def _sensitivity_grid(ctx: ReportContext) -> tuple[list[list[float]], list[list[float | None]]]:
    """Return (van_matrix, tri_matrix) 5x5 for CA x cost factors."""
    van_m: list[list[float]] = []
    tri_m: list[list[float | None]] = []
    base_dump = ctx.inputs.model_dump()
    for ca_f in CA_FACTORS:
        van_row: list[float] = []
        tri_row: list[float | None] = []
        for cost_f in COST_FACTORS:
            data = copy(base_dump)
            ops = data.setdefault("operations", {})
            ops["salePrice"] = float(ops.get("salePrice", 0)) * ca_f
            ops["rawMaterialCost"] = float(ops.get("rawMaterialCost", 0)) * cost_f
            ops["packagingCost"] = float(ops.get("packagingCost", 0)) * cost_f
            inp = PlanInputs.model_validate(data)
            res = calculate_plan(inp, discount_rate=ctx.ind.discountRate)
            van_row.append(res.indicators.van)
            tri_row.append(res.indicators.tri)
        van_m.append(van_row)
        tri_m.append(tri_row)
    return van_m, tri_m


def _sheet_resume(wb: Workbook, ctx: ReportContext) -> None:
    ws = wb.create_sheet("1. Resume executif", 0)
    w = _SheetWriter(ws)
    w.setup_sheet()
    w.cell(1, 1, ctx.title, style="header")
    w.cell(2, 1, ctx.company)
    w.cell(3, 1, f"Reference : {ctx.plan_id[:8].upper() if ctx.plan_id else '—'}")

    fin = ctx.inputs.financing
    loan = fin.loan
    debt = loan.amount or ctx.results.totalInvestment * fin.debtRatio
    equity = ctx.results.totalInvestment * fin.equityRatio

    row = 5
    w.cell(row, 1, "Synthese investissement et financement", style="subtotal")
    row += 1
    recap = [
        ("Investissement total (DT)", ctx.results.totalInvestment),
        ("Fonds propres (DT)", equity),
        ("Credit moyen terme (DT)", debt),
        ("Duree emprunt (ans)", loan.years),
        ("Taux d'interet", loan.rate),
    ]
    for label, val in recap:
        w.cell(row, 1, label, style="label")
        c = w.cell(row, 2, val)
        if isinstance(val, float) and label == "Taux d'interet":
            c.number_format = PCT_FMT
        row += 1

    row += 1
    kpis = [
        ("VAN (DT)", ctx.ind.van, CLR_KPI_VAN),
        ("TRI", ctx.ind.tri or 0, CLR_KPI_VAN),
        ("DRCI (ans)", ctx.ind.drciYears or 0, CLR_KPI_VAN),
        ("CA moyen (DT)", sum(ctx.rev()) / max(1, sum(1 for x in ctx.rev() if x)), CLR_KPI_VAN),
        ("Resultat net moyen (DT)", sum(ctx.net()) / HORIZON, CLR_KPI_VAN),
    ]
    col = 1
    for label, val, clr in kpis:
        w.cell(row, col, label, style="header")
        w.cell(row + 1, col, val)
        w.ws.cell(row, col).fill = _fill(clr)
        w.ws.cell(row + 1, col).fill = _fill(clr)
        if label == "TRI":
            w.ws.cell(row + 1, col).number_format = PCT_FMT
        col += 1

    row += 4
    row = w.write_year_header(row, "Indicateur")
    row = w.write_year_series(row, "Chiffre d'affaires HT", ctx.rev())
    row = w.write_year_series(row, "Resultat net", ctx.net(), row_style="total")

    chart_row = row + 2
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "CA et Resultat net (7 ans)"
    chart.y_axis.title = "DT"
    data = Reference(ws, min_col=2, min_row=row - 2, max_col=1 + HORIZON, max_row=row - 1)
    cats = Reference(ws, min_col=2, min_row=row - 3, max_col=1 + HORIZON)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 10
    ws.add_chart(chart, f"A{chart_row}")
    w.freeze()


def _sheet_investissement(wb: Workbook, ctx: ReportContext) -> None:
    ws = wb.create_sheet("2. Investissement")
    w = _SheetWriter(ws)
    w.setup_sheet()
    r = 1
    w.cell(r, 1, "INVESTISSEMENT ET FINANCEMENT", style="header")
    r += 2
    buckets = _categorize_investments(ctx)
    total_capex_val = ctx.results.totalInvestment or total_capex(ctx.inputs)

    for cat_label, key in [
        ("Immobilisations incorporelles", "incorporel"),
        ("Agencement et amenagement", "agencement"),
        ("Materiel industriel", "industriel"),
        ("Materiel de transport", "transport"),
        ("Mobilier de bureau", "bureau"),
        ("Frais preliminaires", "preliminaires"),
    ]:
        items = buckets.get(key, [])
        w.cell(r, 1, cat_label, style="subtotal")
        r += 1
        w.cell(r, 1, "Designation", style="header")
        w.cell(r, 2, "Valeur DT", style="header")
        w.cell(r, 3, "Taux amort.", style="header")
        w.cell(r, 4, "% du total", style="header")
        r += 1
        sub = 0.0
        for name, amt, life, rate in items:
            pct = _safe_div(amt, total_capex_val)
            w.cell(r, 1, name, style="label")
            w.cell(r, 2, amt)
            w.cell(r, 3, rate)
            w.ws.cell(r, 3).number_format = PCT_FMT
            w.cell(r, 4, pct)
            w.ws.cell(r, 4).number_format = PCT_FMT
            sub += amt
            r += 1
        if not items:
            w.cell(r, 1, "—", style="input")
            w.cell(r, 2, 0)
            r += 1
        w.cell(r, 1, f"Sous-total {cat_label}", style="subtotal")
        w.cell(r, 2, sub, style="subtotal")
        r += 2

    bfr0 = ctx.results.bfr.years[0] if ctx.results.bfr.years else 0
    w.cell(r, 1, "Total investissement", style="total")
    w.cell(r, 2, total_capex_val, style="total")
    r += 1
    w.cell(r, 1, "BFR initial (An 1)", style="label")
    w.cell(r, 2, bfr0)
    r += 1
    w.cell(r, 1, "Investissement + BFR", style="total")
    w.cell(r, 2, total_capex_val + bfr0, style="total")
    r += 2

    w.cell(r, 1, "Schema de financement", style="subtotal")
    r += 1
    fin = ctx.inputs.financing
    eq = total_capex_val * fin.equityRatio
    debt = fin.loan.amount or total_capex_val * fin.debtRatio
    w.cell(r, 1, "Source", style="header")
    w.cell(r, 2, "Valeur DT", style="header")
    w.cell(r, 3, "%", style="header")
    r += 1
    for lbl, val in [
        ("Fonds propres", eq),
        ("Credit moyen terme", debt),
        ("Subventions / aides", 0),
    ]:
        w.cell(r, 1, lbl, style="label")
        w.cell(r, 2, val)
        w.cell(r, 3, _safe_div(val, total_capex_val))
        w.ws.cell(r, 3).number_format = PCT_FMT
        r += 1
    w.cell(r, 1, "Total", style="total")
    w.cell(r, 2, eq + debt, style="total")
    w.cell(r, 3, 1.0, style="total")
    w.ws.cell(r, 3).number_format = PCT_FMT
    w.freeze()


def _sheet_ca(wb: Workbook, ctx: ReportContext) -> None:
    ws = wb.create_sheet("3. Chiffre d affaires")
    w = _SheetWriter(ws)
    w.setup_sheet()
    r = 1
    w.cell(r, 1, "CHIFFRE D'AFFAIRES PREVISIONNEL", style="header")
    r += 2
    products = _product_names(ctx)
    rev = ctx.rev()
    r = w.write_year_header(r, "Produit")
    n = len(products)
    per_prod = [rev[y] / n for y in range(HORIZON)] if n else rev
    for pname in products:
        r = w.write_year_series(r, pname, per_prod)
    growth = []
    for y in range(HORIZON):
        if y == 0:
            growth.append(0.0)
        else:
            growth.append(_safe_div(rev[y] - rev[y - 1], rev[y - 1]))
    r = w.write_year_series(r, "Taux croissance annuel", growth, pct_row=True, row_style="subtotal")
    r = w.write_year_series(r, "CA net total", rev, row_style="total")

    chart = LineChart()
    chart.title = "Evolution du CA (7 ans)"
    data = Reference(ws, min_col=2, min_row=r - 1, max_col=1 + HORIZON)
    cats = Reference(ws, min_col=2, min_row=r - 3, max_col=1 + HORIZON)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 9
    ws.add_chart(chart, f"A{r + 2}")
    w.freeze()


def _sheet_achats(wb: Workbook, ctx: ReportContext) -> None:
    ws = wb.create_sheet("4. Achats consommes")
    w = _SheetWriter(ws)
    w.setup_sheet()
    r = 1
    w.cell(r, 1, "ACHATS CONSOMMES", style="header")
    r += 2
    products = _product_names(ctx)
    mp = ctx.purchase_mp()
    r = w.write_year_header(r, "Matiere premiere / produit")
    n = max(1, len(products))
    for pname in products:
        w.write_year_series(r, pname, [mp[y] / n for y in range(HORIZON)])
        r += 1
    r = w.write_year_series(r, "Sous-total matieres premieres", mp, row_style="subtotal")
    packaging: list[float] = []
    other: list[float] = []
    qty = _series(ctx.results.qtySold, "qty")
    for y in range(HORIZON):
        yr = ctx.yearly[y] if y < len(ctx.yearly) else {}
        cogs = float(yr.get("cogs", 0))
        pack_est = ctx.inputs.operations.packagingCost * (qty[y] if y < len(qty) else 0)
        packaging.append(pack_est)
        other.append(max(0.0, cogs - mp[y] - pack_est))
    r = w.write_year_series(r, "Emballages, aromes, energie", packaging)
    r = w.write_year_series(r, "Autres approvisionnements", other)
    total = [mp[y] + packaging[y] + other[y] for y in range(HORIZON)]
    r = w.write_year_series(r, "Total achats consommes", total, row_style="total")
    w.freeze()


def _sheet_charges(wb: Workbook, ctx: ReportContext) -> None:
    ws = wb.create_sheet("5. Charges exploitation")
    w = _SheetWriter(ws)
    w.setup_sheet()
    r = 1
    w.cell(r, 1, "CHARGES D'EXPLOITATION", style="header")
    r += 2

    personnel_total = sum(p.headcount * p.annualSalary for p in ctx.inputs.plAssumptions.personnel)
    cnss = personnel_total * 0.19
    r = w.write_year_header(r, "Charges de personnel")
    r = w.write_year_series(r, "Salaire brut", [personnel_total] * HORIZON)
    r = w.write_year_series(r, "CNSS et charges sociales (19%)", [cnss] * HORIZON)
    r = w.write_year_series(r, "Total personnel", [personnel_total + cnss] * HORIZON, row_style="subtotal")

    r += 1
    r = w.write_year_header(r, "Effectif par fonction")
    for p in ctx.inputs.plAssumptions.personnel:
        if p.role.strip() or p.headcount:
            r = w.write_year_series(r, p.role, [float(p.headcount)] * HORIZON)
    r += 1

    dep_sched = annual_depreciation_schedule(ctx.inputs)
    r = w.write_year_header(r, "Dotations aux amortissements")
    for eq in ctx.inputs.investments.equipment:
        if eq.cost <= 0:
            continue
        life = max(1, eq.usefulLifeYears)
        annual = eq.cost / life
        dep_line = [0.0] * HORIZON
        start = max(0, min(HORIZON - 1, eq.acquisitionYear - 1))
        for off in range(life):
            yi = start + off
            if yi < HORIZON:
                dep_line[yi] = annual
        r = w.write_year_series(r, eq.name[:40], dep_line)
    r = w.write_year_series(r, "Total DAP", dep_sched, row_style="subtotal")

    r += 1
    r = w.write_year_header(r, "Autres charges d'exploitation")
    r = w.write_year_series(r, "Frais de marketing", ctx.mkt())
    r = w.write_year_series(r, "Transport / distribution", ctx.dist())
    other_line = []
    for y in range(HORIZON):
        yr = ctx.yearly[y] if y < len(ctx.yearly) else {}
        other_line.append(float(yr.get("otherOpex", 0)) + float(yr.get("vat", 0)))
    r = w.write_year_series(r, "Autres charges", other_line)
    tot = [ctx.mkt()[y] + ctx.dist()[y] + other_line[y] for y in range(HORIZON)]
    r = w.write_year_series(r, "Total autres charges", tot, row_style="total")
    w.freeze()


def _sheet_emprunt(wb: Workbook, ctx: ReportContext) -> None:
    ws = wb.create_sheet("6. Tableau amortissement")
    w = _SheetWriter(ws)
    w.setup_sheet()
    fin = ctx.inputs.financing
    loan = fin.loan
    r = 1
    w.cell(r, 1, "TABLEAU D'AMORTISSEMENT EMPRUNT", style="header")
    r += 2
    w.cell(r, 1, f"Taux : {loan.rate:.2%} — Duree : {loan.years} ans — Differe principal : {loan.graceMonthsPrincipal} mois")
    r += 2
    headers = ["Poste", *YEAR_LABELS, "Total"]
    for i, h in enumerate(headers):
        w.cell(r, 1 + i, h, style="header")
    r += 1
    loan_rows = _loan_annual_rows(ctx)
    lines = [
        ("Capital restant du (debut)", [x["opening"] for x in loan_rows]),
        ("Interets", [x["interest"] for x in loan_rows]),
        ("Remboursement principal", [x["principal"] for x in loan_rows]),
        ("Service de la dette", [x["service"] for x in loan_rows]),
        ("Capital restant du (fin)", [x["closing"] for x in loan_rows]),
    ]
    for label, vals in lines:
        w.cell(r, 1, label, style="label")
        for i, v in enumerate(vals):
            w.cell(r, 2 + i, v)
        w.cell(r, 2 + HORIZON, sum(vals), style="subtotal")
        r += 1
    w.cell(r, 1, "Total interets payes", style="total")
    w.cell(r, 2 + HORIZON, sum(x["interest"] for x in loan_rows), style="total")
    w.freeze()


def _sheet_pl(wb: Workbook, ctx: ReportContext) -> None:
    ws = wb.create_sheet("7. Etat resultat")
    w = _SheetWriter(ws)
    w.setup_sheet()
    r = 1
    w.cell(r, 1, "ETAT DE RESULTAT PREVISIONNEL", style="header")
    r += 2
    w.cell(r, 1, "Poste", style="header")
    w.cell(r, 2, "Total 7 ans", style="header")
    w.cell(r, 3, "% Revenu", style="header")
    r += 1
    rev_total = sum(ctx.rev()) or 1.0
    for label, key, style in [
        ("Chiffre d'affaires HT", "revenue", CLR_PROD),
        ("Achats consommes", "cogs", CLR_CHARGE),
        ("Charges de personnel", "personnel", CLR_CHARGE),
        ("Autres charges exploitation", "otherOpex", CLR_CHARGE),
        ("Frais distribution", "distribution", CLR_CHARGE),
        ("Frais marketing", "marketing", CLR_CHARGE),
        ("TVA nette", "vat", CLR_CHARGE),
        ("EBE", "ebe", CLR_RESULT),
        ("Dotations amortissement", "depreciation", CLR_CHARGE),
        ("Resultat exploitation (EBIT)", None, CLR_RESULT),
        ("Charges financieres", "interest", CLR_CHARGE),
        ("Resultat avant impot", None, CLR_RESULT),
        ("Impot sur societes", "tax", CLR_CHARGE),
        ("Resultat net", "netProfit", CLR_RESULT),
    ]:
        if key:
            total = sum(float(ctx.yearly[y].get(key, 0)) for y in range(min(HORIZON, len(ctx.yearly))))
        elif label == "Resultat exploitation (EBIT)":
            total = sum(
                float(ctx.yearly[y].get("ebe", 0)) - float(ctx.yearly[y].get("depreciation", 0))
                for y in range(min(HORIZON, len(ctx.yearly)))
            )
        elif label == "Resultat avant impot":
            total = sum(
                float(ctx.yearly[y].get("ebe", 0))
                - float(ctx.yearly[y].get("depreciation", 0))
                - float(ctx.yearly[y].get("interest", 0))
                for y in range(min(HORIZON, len(ctx.yearly)))
            )
        else:
            total = sum(ctx.net())
        pct_rev = _safe_div(total, rev_total)
        w.cell(r, 1, label, style="label")
        w.ws.cell(r, 1).fill = _fill(style)
        w.cell(r, 2, total)
        w.cell(r, 3, pct_rev)
        w.ws.cell(r, 3).number_format = PCT_FMT
        r += 1

    chart = LineChart()
    chart.title = "Resultat exploitation et net"
    r += 1
    w.freeze()


def _sheet_cashflows(wb: Workbook, ctx: ReportContext) -> None:
    ws = wb.create_sheet("8. Cash-flows")
    w = _SheetWriter(ws)
    w.setup_sheet()
    r = 1
    w.cell(r, 1, "CASH-FLOWS ET RENTABILITE", style="header")
    r += 2
    disc = ctx.ind.discountRate
    net = ctx.net()
    dep = ctx.dep()
    bfr_v = ctx.bfr_var()
    ocf = [net[y] + dep[y] - bfr_v[y] for y in range(HORIZON)]
    cum = []
    s = 0.0
    for v in ocf:
        s += v
        cum.append(s)
    discounted = [v / ((1 + disc) ** (y + 1)) for y, v in enumerate(ocf)]
    cum_disc = []
    s = 0.0
    for v in discounted:
        s += v
        cum_disc.append(s)

    r = w.write_year_header(r, "Poste")
    r = w.write_year_series(r, "Resultat net", net)
    r = w.write_year_series(r, "Dotations aux amortissements", dep)
    r = w.write_year_series(r, "Variation BFR", [-v for v in bfr_v])
    r = w.write_year_series(r, "Cash-flow exploitation", ocf, row_style="subtotal")
    r = w.write_year_series(r, "Cash-flows cumules", cum)
    r = w.write_year_series(r, "Cash-flows actualises", discounted)
    r = w.write_year_series(r, "CF actualises cumules", cum_disc)

    inv = ctx.results.totalInvestment
    r += 1
    w.cell(r, 1, "VCN immobilisations + recuperation BFR (An 7)", style="label")
    salvage = inv * 0.1 + (ctx.results.bfr.years[-1] if ctx.results.bfr.years else 0)
    w.cell(r, 2 + HORIZON - 1, salvage)
    r += 2

    cf_row = r
    w.cell(r, 1, "Flux pour VAN (t=0 investissement)", style="label")
    w.cell(r, 2, -inv, style="input")
    for y in range(HORIZON):
        w.cell(r, 3 + y, ocf[y])
    r += 1
    van_cell = f"{get_column_letter(3)}{r}"
    last_cf_col = get_column_letter(2 + HORIZON)
    w.cell(r, 1, "VAN (formule Excel)", style="total")
    w.cell(r, 2, f"=NPV({disc},{get_column_letter(3)}{cf_row}:{last_cf_col}{cf_row})+B{cf_row}")
    w.ws.cell(r, 2).number_format = NUM_FMT
    r += 1
    w.cell(r, 1, "TRI (formule Excel)", style="total")
    w.cell(r, 2, f"=IRR(B{cf_row}:{last_cf_col}{cf_row})")
    w.ws.cell(r, 2).number_format = PCT_FMT
    r += 1
    w.cell(r, 1, "DRCI (ans)", style="label")
    drci = ctx.ind.drciYears
    w.cell(r, 2, drci if drci else "—")
    if drci and 1 <= int(round(drci)) <= HORIZON:
        col = 2 + int(round(drci)) - 1
        w.ws.cell(cf_row - 2, col).fill = _fill(CLR_DRCI)

    area = AreaChart()
    area.title = "Cash-flows cumules"
    area.add_data(Reference(ws, min_col=2, min_row=r - 6, max_col=1 + HORIZON), titles_from_data=True)
    ws.add_chart(area, f"A{r + 2}")
    w.freeze()


def _sheet_sensibilite(wb: Workbook, ctx: ReportContext) -> None:
    ws = wb.create_sheet("9. Sensibilite")
    w = _SheetWriter(ws)
    w.setup_sheet()
    van_m, tri_m = _sensitivity_grid(ctx)
    r = 1
    w.cell(r, 1, "ANALYSE DE SENSIBILITE", style="header")
    r += 2
    for title, matrix, fmt in [
        ("VAN selon variation CA (colonnes) et couts (lignes)", van_m, NUM_FMT),
        ("TRI selon variation CA et couts", tri_m, PCT_FMT),
    ]:
        w.cell(r, 1, title, style="subtotal")
        r += 1
        w.cell(r, 1, "Couts \\ CA", style="header")
        for j, ca in enumerate(CA_FACTORS):
            w.cell(r, 2 + j, f"{ca:.0%}", style="header")
        r += 1
        for i, cost_f in enumerate(COST_FACTORS):
            w.cell(r, 1, f"{cost_f:.0%}", style="label")
            for j in range(5):
                val = matrix[i][j]
                c = w.cell(r, 2 + j, val if val is not None else "—")
                c.number_format = fmt
            r += 1
        r += 2

    # Conditional formatting on VAN block (rows 4-8 cols B-F approx)
    ws.conditional_formatting.add(
        "B4:F8",
        CellIsRule(operator="lessThan", formula=["0"], fill=_fill("FFCDD2")),
    )
    ws.conditional_formatting.add(
        "B4:F8",
        CellIsRule(operator="greaterThan", formula=["100000"], fill=_fill("C8E6C9")),
    )
    w.freeze()


def _sheet_ratios(wb: Workbook, ctx: ReportContext) -> None:
    ws = wb.create_sheet("10. Marge et ratios")
    w = _SheetWriter(ws)
    w.setup_sheet()
    r = 1
    w.cell(r, 1, "MARGE BRUTE ET RATIOS", style="header")
    r += 2
    rev = ctx.rev()
    margin = []
    roi = []
    coverage = []
    breakeven = []
    inv = ctx.results.totalInvestment or 1
    for y in range(HORIZON):
        yr = ctx.yearly[y] if y < len(ctx.yearly) else {}
        rev_y = rev[y]
        margin.append(float(yr.get("grossMarginPct", _safe_div(rev_y - float(yr.get("cogs", 0)), rev_y))))
        roi.append(_safe_div(ctx.net()[y], inv))
        debt_svc = ctx.interest()[y] + ctx.principal()[y]
        coverage.append(_safe_div(ctx.ocf()[y], debt_svc) if debt_svc else 0)
        m_rate = margin[-1]
        fixed = float(yr.get("personnel", 0)) + float(yr.get("otherOpex", 0)) + float(yr.get("depreciation", 0))
        breakeven.append(_safe_div(fixed, m_rate) if m_rate else 0)
    r = w.write_year_header(r, "Indicateur")
    r = w.write_year_series(r, "Taux marge brute", margin, pct_row=True)
    r = w.write_year_series(r, "ROI (RN / investissement)", roi, pct_row=True)
    r = w.write_year_series(r, "Couverture service dette", coverage)
    r = w.write_year_series(r, "Point mort (CA seuil)", breakeven)
    w.freeze()


def generate_excel_report(plan_data: dict, output_path: str) -> None:
    """
    Build the full VIPA-style Excel workbook at output_path.

    plan_data: dict with keys inputs, results, optional yearly, title, plan_id.
    """
    ctx = _parse_plan_data(plan_data)
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_resume(wb, ctx)
    _sheet_investissement(wb, ctx)
    _sheet_ca(wb, ctx)
    _sheet_achats(wb, ctx)
    _sheet_charges(wb, ctx)
    _sheet_emprunt(wb, ctx)
    _sheet_pl(wb, ctx)
    _sheet_cashflows(wb, ctx)
    _sheet_sensibilite(wb, ctx)
    _sheet_ratios(wb, ctx)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def celery_export_excel(plan_data: dict, output_path: str) -> str:
    """Alias for Celery task integration; returns absolute path."""
    generate_excel_report(plan_data, output_path)
    return str(Path(output_path).resolve())
