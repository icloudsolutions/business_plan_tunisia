"""Rich PDF and Excel export builders for validated business plans."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from bp_schema.liasse import PlanInputs, PlanResults
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

HORIZON = 7
_YEAR_HEADERS = [f"An {i + 1}" for i in range(HORIZON)]


def _fmt(n: float | None, *, digits: int = 0) -> str:
    if n is None:
        return "—"
    return f"{n:,.{digits}f}".replace(",", " ")


def _pct(n: float | None) -> str:
    if n is None:
        return "—"
    return f"{n * 100:.2f}%"


def _year_val(series, y: int) -> float:
    years = getattr(series, "years", series) if not isinstance(series, list) else series
    if y < len(years):
        return float(years[y])
    return 0.0


def _pdf_safe(text: str) -> str:
    return str(text).encode("latin-1", errors="replace").decode("latin-1")


def _investment_rows(inputs: PlanInputs) -> list[list[str]]:
    rows: list[list[str]] = []
    for eq in inputs.investments.equipment:
        if eq.cost > 0 or eq.name.strip():
            rows.append(
                [
                    eq.name,
                    "Incorporel" if eq.assetType == "intangible" else "Corporel",
                    _fmt(eq.cost),
                    str(eq.usefulLifeYears),
                    f"An {eq.acquisitionYear}",
                ]
            )
    for line in inputs.investments.intangible:
        if line.amount > 0:
            rows.append(
                [line.label, "Incorporel", _fmt(line.amount), str(line.usefulLifeYears), "—"]
            )
    for line in inputs.investments.tangible:
        if line.amount > 0:
            rows.append(
                [line.label, "Corporel", _fmt(line.amount), str(line.usefulLifeYears), "—"]
            )
    return rows


def _personnel_rows(inputs: PlanInputs) -> list[list[str]]:
    return [
        [p.role, str(p.headcount), _fmt(p.annualSalary)]
        for p in inputs.plAssumptions.personnel
        if p.role.strip() or p.headcount or p.annualSalary
    ]


def _pl_metric_rows(results: PlanResults) -> list[tuple[str, list[float]]]:
    return [
        ("Chiffre d'affaires HT (TND)", [_year_val(results.revenue, y) for y in range(HORIZON)]),
        ("Résultat net (TND)", [_year_val(results.netProfit, y) for y in range(HORIZON)]),
        (
            "Cash-flow exploitation (TND)",
            [_year_val(results.operatingCashFlow, y) for y in range(HORIZON)],
        ),
        (
            "Trésorerie cumulée (TND)",
            [_year_val(results.cumulativeTreasury, y) for y in range(HORIZON)],
        ),
        ("BFR (TND)", [_year_val(results.bfr, y) for y in range(HORIZON)]),
        ("Variation BFR (TND)", [_year_val(results.bfrVariation, y) for y in range(HORIZON)]),
        ("Amortissements (TND)", [_year_val(results.depreciation, y) for y in range(HORIZON)]),
        ("Intérêts (TND)", [_year_val(results.interestExpense, y) for y in range(HORIZON)]),
        (
            "Remboursement principal (TND)",
            [_year_val(results.principalRepayment, y) for y in range(HORIZON)],
        ),
    ]


def _xlsx_style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _xlsx_autowidth(ws, min_width: int = 10, max_width: int = 42) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_width
        for cell in col:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = width


def _xlsx_write_kv_sheet(ws, title: str, rows: list[tuple[str, str]]) -> None:
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Rubrique", "Valeur"])
    _xlsx_style_header(ws, row=3)
    for label, value in rows:
        ws.append([label, value])
    _xlsx_autowidth(ws)


def _xlsx_write_table(ws, headers: list, rows: list[list]) -> int:
    ws.append(headers)
    header_row = ws.max_row
    _xlsx_style_header(ws, row=header_row)
    for row in rows:
        ws.append(row)
    _xlsx_autowidth(ws)
    return header_row


def build_export_xlsx(
    plan_id: str,
    inputs: PlanInputs,
    results: PlanResults,
    path: Path,
    *,
    plan_title: str | None = None,
) -> str:
    company = inputs.company.name.strip() or "Projet"
    ind = results.indicators
    ops = inputs.operations
    fin = inputs.financing
    wc = inputs.workingCapital
    loan = fin.loan

    wb = Workbook()
    # —— Synthèse ——
    ws0 = wb.active
    ws0.title = "Synthèse"
    _xlsx_write_kv_sheet(
        ws0,
        plan_title or f"Business Plan — {company}",
        [
            ("Société", company),
            ("Forme juridique", inputs.company.legalForm),
            ("Référence", plan_id[:8].upper()),
            ("Date export", datetime.now().strftime("%d/%m/%Y %H:%M")),
            ("Investissement total (TND)", _fmt(results.totalInvestment)),
            ("VAN (TND)", _fmt(ind.van)),
            ("TRI", _pct(ind.tri)),
            ("DRCI (ans)", _fmt(ind.drciYears, digits=1) if ind.drciYears else "—"),
            ("Taux actualisation", _pct(ind.discountRate)),
            ("Bilan équilibré", "Oui" if results.balanceSheetBalanced else "Non"),
            ("BFR cohérent", "Oui" if results.bfrCoherent else "Non"),
            (
                "Trésorerie",
                "Positive sur 7 ans"
                if results.cashRunwayBreakYear is None
                else f"Rupture an {results.cashRunwayBreakYear}",
            ),
            ("Jours ouvrés / an", _fmt(ops.workingDaysPerYear, digits=0)),
            ("Prix de vente unitaire (TND)", _fmt(ops.salePrice)),
            ("Fonds propres", _pct(fin.equityRatio)),
            ("Dette", _pct(fin.debtRatio)),
            ("Taux emprunt", _pct(loan.rate)),
            ("Durée emprunt (ans)", str(loan.years)),
        ],
    )

    # —— Investissements ——
    ws1 = wb.create_sheet("Investissements")
    inv = _investment_rows(inputs)
    if inv:
        _xlsx_write_table(
            ws1,
            ["Désignation", "Nature", "Montant TND", "Amort. ans", "Mise en service"],
            inv,
        )
    else:
        ws1.append(["Aucun investissement détaillé"])
    ws1.append([])
    ws1.append(["Total CAPEX (TND)", _fmt(results.totalInvestment)])

    # —— Hypothèses ——
    ws2 = wb.create_sheet("Hypothèses")
    _xlsx_write_kv_sheet(
        ws2,
        "Hypothèses d'exploitation et de financement",
        [
            ("Heures / jour", _fmt(ops.hoursPerDay, digits=1)),
            ("Coût matière unitaire", _fmt(ops.rawMaterialCost)),
            ("Coût conditionnement", _fmt(ops.packagingCost)),
            ("Taux déchet", _pct(ops.wasteRate.value)),
            ("Délai clients (j)", str(wc.clientPaymentDays)),
            ("Délai fournisseurs (j)", str(wc.supplierPaymentDays)),
            ("Stock PF (j)", str(wc.finishedGoodsStockDays)),
            ("Stock MP (mois)", _fmt(wc.rawMaterialStockMonths, digits=1)),
            ("Montant emprunt (TND)", _fmt(loan.amount or results.totalInvestment * fin.debtRatio)),
            ("Différé principal (mois)", str(loan.graceMonthsPrincipal)),
            ("IS", _pct(inputs.plAssumptions.corporateTaxRate)),
            ("Frais distribution % CA", _pct(inputs.plAssumptions.distributionExpensePct)),
            ("Frais marketing % CA", _pct(inputs.plAssumptions.marketingExpensePct)),
        ],
    )

    # —— P&L 7 ans ——
    ws3 = wb.create_sheet("Compte resultat 7 ans")
    pl_rows = [[label, *vals] for label, vals in _pl_metric_rows(results)]
    _xlsx_write_table(ws3, ["Poste", *_YEAR_HEADERS], pl_rows)
    for row in ws3.iter_rows(min_row=4, max_row=ws3.max_row, min_col=2, max_col=1 + HORIZON):
        for cell in row:
            cell.number_format = '#,##0'

    # —— Trésorerie détail ——
    ws4 = wb.create_sheet("Tresorerie BFR")
    treas_rows = [
        [
            "Trésorerie cumulée",
            *[_fmt(_year_val(results.cumulativeTreasury, y)) for y in range(HORIZON)],
        ],
        ["BFR", *[_fmt(_year_val(results.bfr, y)) for y in range(HORIZON)]],
        [
            "Variation BFR",
            *[_fmt(_year_val(results.bfrVariation, y)) for y in range(HORIZON)],
        ],
    ]
    _xlsx_write_table(ws4, ["Poste", *_YEAR_HEADERS], treas_rows)

    # —— Personnel ——
    pers = _personnel_rows(inputs)
    if pers:
        ws5 = wb.create_sheet("Personnel")
        _xlsx_write_table(ws5, ["Poste", "Effectif", "Salaire annuel TND"], pers)

    # —— Volumes (if present) ——
    if any(_year_val(results.qtySold, y) > 0 for y in range(HORIZON)):
        ws6 = wb.create_sheet("Volumes")
        vol_rows = [
            ["Quantités vendues", *[_fmt(_year_val(results.qtySold, y), digits=0) for y in range(HORIZON)]],
            [
                "Quantités produites",
                *[_fmt(_year_val(results.qtyProduced, y), digits=0) for y in range(HORIZON)],
            ],
            [
                "Stock PF clôture",
                *[_fmt(_year_val(results.closingStockPF, y), digits=0) for y in range(HORIZON)],
            ],
        ]
        _xlsx_write_table(ws6, ["Indicateur", *_YEAR_HEADERS], vol_rows)

    out = path / f"business_plan_{plan_id}.xlsx"
    wb.save(out)
    return str(out.resolve())


def _pdf_table(data: list[list], col_widths: list[float] | None = None) -> Table:
    t = Table(data, colWidths=col_widths)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return t


def build_export_pdf(
    plan_id: str,
    inputs: PlanInputs,
    results: PlanResults,
    path: Path,
    *,
    plan_title: str | None = None,
) -> str:
    company = inputs.company.name.strip() or "Sans nom"
    ind = results.indicators
    out_path = path / f"business_plan_{plan_id}.pdf"

    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=A4,
        leftMargin=1.8 * cm,
        rightMargin=1.8 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=12,
        textColor=colors.HexColor("#1E3A5F"),
    )
    h2 = ParagraphStyle(
        "H2",
        parent=styles["Heading2"],
        fontSize=12,
        spaceBefore=14,
        spaceAfter=8,
        textColor=colors.HexColor("#1E3A5F"),
    )
    body = styles["Normal"]

    story: list = []

    story.append(
        Paragraph(
            _pdf_safe(plan_title or f"Business Plan — {company}"),
            title_style,
        )
    )
    story.append(
        Paragraph(
            _pdf_safe(
                f"Etude de faisabilite | {inputs.company.legalForm} | "
                f"{datetime.now().strftime('%d/%m/%Y')} | Ref. {plan_id[:8].upper()}"
            ),
            body,
        )
    )
    story.append(Spacer(1, 12))

    story.append(Paragraph(_pdf_safe("1. Synthese financiere"), h2))
    summary_data = [
        ["Indicateur", "Valeur"],
        ["Investissement total (TND)", _fmt(results.totalInvestment)],
        ["VAN (TND)", _fmt(ind.van)],
        ["TRI", _pct(ind.tri)],
        ["DRCI (ans)", _fmt(ind.drciYears, digits=1) if ind.drciYears else "—"],
        ["Bilan equilibre", "Oui" if results.balanceSheetBalanced else "Non"],
        ["BFR coherent", "Oui" if results.bfrCoherent else "Non"],
        [
            "Tresorerie 7 ans",
            "OK"
            if results.cashRunwayBreakYear is None
            else f"Alerte an {results.cashRunwayBreakYear}",
        ],
    ]
    story.append(_pdf_table(summary_data, [8 * cm, 8 * cm]))
    story.append(Spacer(1, 10))

    inv = _investment_rows(inputs)
    if inv:
        story.append(Paragraph(_pdf_safe("2. Programme d'investissement"), h2))
        inv_data = [
            ["Designation", "Nature", "TND", "Amort.", "An"],
            *[
                [r[0], r[1], r[2], r[3], r[4]]
                for r in inv[:12]
            ],
        ]
        if len(inv) > 12:
            inv_data.append(["...", f"{len(inv) - 12} lignes supplementaires", "", "", ""])
        story.append(
            _pdf_table(inv_data, [5 * cm, 2.5 * cm, 2.5 * cm, 1.5 * cm, 1.5 * cm])
        )
        story.append(Spacer(1, 10))

    story.append(Paragraph(_pdf_safe("3. Compte de resultat et tresorerie (7 ans)"), h2))
    pl_data: list[list] = [["Poste"] + _YEAR_HEADERS]
    for label, vals in _pl_metric_rows(results)[:8]:
        pl_data.append([label] + [_fmt(v) for v in vals])
    col_w = [4.2 * cm] + [1.85 * cm] * HORIZON
    story.append(_pdf_table(pl_data, col_w))

    fin = inputs.financing
    story.append(Spacer(1, 10))
    story.append(Paragraph(_pdf_safe("4. Financement et BFR"), h2))
    fin_data = [
        ["Poste", "Valeur"],
        ["Fonds propres", _pct(fin.equityRatio)],
        ["Dette", _pct(fin.debtRatio)],
        ["Taux emprunt", _pct(fin.loan.rate)],
        ["Delai clients (j)", str(inputs.workingCapital.clientPaymentDays)],
        ["Delai fournisseurs (j)", str(inputs.workingCapital.supplierPaymentDays)],
    ]
    story.append(_pdf_table(fin_data, [8 * cm, 8 * cm]))

    pers = _personnel_rows(inputs)
    if pers:
        story.append(Spacer(1, 10))
        story.append(Paragraph(_pdf_safe("5. Masse salariale"), h2))
        pers_data = [["Poste", "Effectif", "Salaire TND"]] + pers[:10]
        story.append(_pdf_table(pers_data, [6 * cm, 2 * cm, 4 * cm]))

    story.append(Spacer(1, 12))
    conclusion = (
        f"Projet « {company} » : "
        + (
            "indicateurs favorables sous hypotheses retenues."
            if ind.van >= 0 and results.cashRunwayBreakYear is None
            else "points de vigilance (VAN, tresorerie ou structure) — ajuster avant depot."
        )
    )
    story.append(Paragraph(_pdf_safe(conclusion), body))

    doc.build(story)
    return str(out_path.resolve())
